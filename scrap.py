from bs4 import BeautifulSoup
import requests
import openpyxl

baseUrl = "https://url.com.br/"

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.157 Safari/537.36' 
}

r = requests.get(baseUrl)

soup = BeautifulSoup(r.content, 'lxml')

productList = soup.find_all('div', class_="product")

productLinks = []

for item in productList:
    for link in item.find_all('a', href=True):
        link = productLinks.append(link['href'])

nomes = []
precos = []

for link in productLinks:
    # print(link)

    r = requests.get(link, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')

    nomes.append(soup.find('div', class_="name").text)
    precos.append(soup.find('div', class_="price").text)

wb = openpyxl.Workbook()

sheet = wb.active
sheet.title = 'Produtos e Valor'

sheet.cell(row=1, column=1).value = 'Produto'
sheet.cell(row=1, column=2).value = 'Valor'

for i in range(0, len(nomes)):
    sheet.cell(row=i+2, column=1).value = nomes[i]
    sheet.cell(row=i+2, column=2).value = precos[i]

wb.save('produtos.xlsx')
